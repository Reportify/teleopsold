from typing import List
from django.db import transaction
from django.db.models import Q
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from core.permissions.tenant_permissions import TenantScopedPermission
from .models import EquipmentInventoryItem, TechnologyTag
from .serializers import (
    EquipmentInventoryItemSerializer,
    TechnologyTagSerializer,
)


class TechnologyTagViewSet(viewsets.ModelViewSet):
    serializer_class = TechnologyTagSerializer
    permission_classes = [IsAuthenticated, TenantScopedPermission]

    def get_queryset(self):
        tenant = getattr(self.request, 'tenant', None)
        qs = TechnologyTag.objects.all()
        if tenant:
            qs = qs.filter(tenant=tenant)
        return qs.order_by('name')

    def perform_create(self, serializer):
        tenant = getattr(self.request, 'tenant', None)
        serializer.save(tenant=tenant, created_by=self.request.user)


class EquipmentInventoryItemViewSet(viewsets.ModelViewSet):
    serializer_class = EquipmentInventoryItemSerializer
    permission_classes = [IsAuthenticated, TenantScopedPermission]

    def get_queryset(self):
        tenant = getattr(self.request, 'tenant', None)
        qs = EquipmentInventoryItem.objects.all()
        if tenant:
            qs = qs.filter(tenant=tenant)

        # Filters
        search = self.request.query_params.get('search')
        category = self.request.query_params.get('category')
        manufacturer = self.request.query_params.get('manufacturer')
        technology = self.request.query_params.get('technology')

        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(material_code__icontains=search)
                | Q(manufacturer__icontains=search)
                | Q(category__icontains=search)
                | Q(sub_category__icontains=search)
            )

        if category:
            qs = qs.filter(category__iexact=category)
        if manufacturer:
            qs = qs.filter(manufacturer__icontains=manufacturer)
        if technology:
            qs = qs.filter(technologies__name__iexact=technology)

        return qs.select_related('tenant').prefetch_related('technologies').order_by('name')

    def perform_create(self, serializer):
        tenant = getattr(self.request, 'tenant', None)
        serializer.save(tenant=tenant, created_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='categories')
    def categories(self, request):
        """Return distinct categories for current tenant with optional search.
        Query params:
          - search: optional text to filter category names (icontains)
        Response: { "categories": ["CABINET", "ANTENNA", ...] }
        """
        tenant = getattr(request, 'tenant', None)
        qs = EquipmentInventoryItem.objects.all()
        if tenant:
            qs = qs.filter(tenant=tenant)

        search = request.query_params.get('search')
        if search:
            qs = qs.filter(category__icontains=search)

        values = (
            qs.exclude(category__isnull=True)
              .exclude(category__exact='')
              .values_list('category', flat=True)
              .distinct()
              .order_by('category')
        )
        return Response({ 'categories': list(values) })

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        tenant = getattr(request, 'tenant', None)
        created = updated = skipped = 0
        errors: List[str] = []
        dry_run = str(request.data.get('dry_run', 'false')).lower() == 'true'
        all_or_nothing = str(request.data.get('all_or_nothing', 'false')).lower() == 'true'

        def normalize_uom(value: str) -> str:
            mapping = {
                'EA': 'Unit',
                'UNIT': 'Unit',
                'UNITS': 'Unit',
                'M': 'Meter',
                'KG': 'Kilogram',
                'L': 'Liter',
                'SQM': 'Square Meter',
            }
            v = (value or '').strip()
            upper = v.upper()
            return mapping.get(upper, 'Unit' if not v else v)

        def normalize_code(raw_code: object, name_value: str) -> str:
            """Sanitize material code:
            - Trim
            - Treat NA/None/- as empty
            - If equals name (case-insensitive), drop it
            - Uppercase and collapse spaces
            """
            if raw_code is None:
                return ''
            code = str(raw_code).strip()
            if not code:
                return ''
            if code.lower() in {'na', 'n/a', 'null', 'none', '-'}:
                return ''
            if name_value:
                # Compare on sanitized tokens (remove non-alnum, collapse spaces)
                import re
                def sanitize(s: str) -> str:
                    return re.sub(r'[^A-Za-z0-9]', '', s).lower()
                name_s = sanitize(name_value)
                code_s = sanitize(code)
                # Cases to drop: equals name, equals name + '1'
                if code_s == name_s or code_s == f"{name_s}1":
                    return ''
            # canonical form
            code = code.replace(' ', '')
            return code.upper()

        rows: List[dict] = []

        def queue_row(obj):
            nonlocal created, updated, skipped
            try:
                name_value = obj.get('name') or obj.get('model_name') or 'Unnamed'
                material_code = normalize_code(obj.get('material_code'), name_value)
                rows.append({
                    'name': name_value,
                    'material_code': material_code,
                    'description': obj.get('description', ''),
                    'category': (obj.get('category') or '').strip(),
                    'sub_category': (obj.get('sub_category') or '').strip(),
                    'manufacturer': (obj.get('manufacturer') or '').strip(),
                    'unit_of_measurement': normalize_uom(obj.get('unit_of_measurement') or obj.get('uom') or ''),
                    'technologies': obj.get('technologies') or obj.get('technology') or obj.get('Technology') or [],
                    'is_active': obj.get('is_active', True),
                })
            except Exception as e:
                errors.append(str(e))

        items = request.data.get('items')
        if isinstance(items, list):
            for obj in items:
                queue_row(obj)
            # fallthrough to write phase below

        upload = request.FILES.get('file')
        if upload:
            import os
            import csv
            from io import TextIOWrapper
            ext = os.path.splitext(upload.name)[1].lower()

            if ext in ['.xlsx', '.xls']:
                try:
                    import openpyxl
                except Exception:
                    return Response({'detail': 'openpyxl is required for Excel uploads'}, status=400)
                wb = openpyxl.load_workbook(upload, data_only=True)

                # Sheet selection and header detection
                requested_sheet = request.data.get('sheet')
                ws = None
                headers: List[str] = []
                header_row_index = 1

                def detect_headers(sheet, start: int = 1, scan: int = 10):
                    for r in range(start, start + scan):
                        row = list(sheet.iter_rows(min_row=r, max_row=r, values_only=True))
                        if not row:
                            continue
                        values = [str(v).strip() if v is not None else '' for v in row[0]]
                        if any(values):
                            return values, r
                    return [], start

                if requested_sheet and requested_sheet in wb.sheetnames:
                    ws = wb[requested_sheet]
                    headers, header_row_index = detect_headers(ws)
                if ws is None:
                    # Find a sheet that contains required columns
                    for sheet in wb.worksheets:
                        hdrs, hrow = detect_headers(sheet)
                        s = {h.lower() for h in hdrs}
                        if {'material code', 'material description'}.issubset(s):
                            ws, headers, header_row_index = sheet, hdrs, hrow
                            break
                    if ws is None:
                        ws = wb.active
                        headers, header_row_index = detect_headers(ws)

                for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
                    row_dict = {headers[i].strip(): (row[i] if (row and i < len(row)) else '') for i in range(len(headers))}
                    obj = {
                        'material_code': row_dict.get('Material Code') or row_dict.get('material_code') or row_dict.get('MaterialCode') or '',
                        'manufacturer': row_dict.get('Manufacturer') or row_dict.get('manufacturer') or '',
                        'name': row_dict.get('Material Description') or row_dict.get('name') or row_dict.get('Model') or '',
                        'category': row_dict.get('Material Category') or row_dict.get('Category') or '',
                        'sub_category': row_dict.get('Material Sub-Category') or row_dict.get('Sub Category') or row_dict.get('Sub-Category') or '',
                        'uom': row_dict.get('UOM') or row_dict.get('uom') or '',
                        'technology': row_dict.get('Technology') or row_dict.get('Technonoly') or row_dict.get('Technologies') or '',
                    }
                    queue_row(obj)
            elif ext in ['.csv', '.txt']:
                reader = csv.DictReader(TextIOWrapper(upload.file, encoding='utf-8'))
                for row in reader:
                    obj = {
                        'material_code': row.get('Material Code') or row.get('material_code') or row.get('MaterialCode') or '',
                        'manufacturer': row.get('Manufacturer') or row.get('manufacturer') or '',
                        'name': row.get('Material Description') or row.get('name') or row.get('Model') or '',
                        'category': row.get('Material Category') or row.get('Category') or '',
                        'sub_category': row.get('Material Sub-Category') or row.get('Sub Category') or row.get('Sub-Category') or '',
                        'uom': row.get('UOM') or row.get('uom') or '',
                        'technology': row.get('Technology') or row.get('Technonoly') or row.get('Technologies') or '',
                    }
                    queue_row(obj)
            else:
                return Response({'detail': 'Unsupported file type. Use .xlsx or .csv'}, status=400)

            # Write phase
            if errors and all_or_nothing:
                return Response({'created': 0, 'updated': 0, 'skipped': len(rows), 'errors': errors, 'sheet_used': (ws.title if 'ws' in locals() and ws is not None else None)}, status=400)

            if dry_run:
                return Response({'created': 0, 'updated': 0, 'skipped': 0, 'errors': errors, 'sheet_used': (ws.title if 'ws' in locals() and ws is not None else None), 'processed_rows': len(rows)})

            # Prefetch existing by material code (non-empty)
            material_codes = [r['material_code'] for r in rows if r['material_code']]
            existing_by_code = {}
            if material_codes:
                for itm in EquipmentInventoryItem.objects.filter(tenant=tenant, material_code__in=material_codes):
                    existing_by_code[itm.material_code] = itm

            # Technology tags prefetch/create mapping
            tech_names = set()
            for r in rows:
                tv = r['technologies']
                if isinstance(tv, str):
                    tech_names.update([t.strip() for t in tv.replace(';', ',').split(',') if t.strip()])
                elif isinstance(tv, list):
                    tech_names.update([str(t).strip() for t in tv if str(t).strip()])
            name_to_tag = {}
            for name in tech_names:
                tag, _ = TechnologyTag.objects.get_or_create(tenant=tenant, name=name, defaults={'created_by': request.user})
                name_to_tag[name] = tag

            # Build create/update lists
            to_create = []
            to_update = []
            row_to_item = {}
            for r in rows:
                tech_vals = r['technologies']
                if isinstance(tech_vals, str):
                    tech_list = [t.strip() for t in tech_vals.replace(';', ',').split(',') if t.strip()]
                elif isinstance(tech_vals, list):
                    tech_list = [str(t).strip() for t in tech_vals if str(t).strip()]
                else:
                    tech_list = []
                r['technologies'] = tech_list

                if r['material_code'] and r['material_code'] in existing_by_code:
                    itm = existing_by_code[r['material_code']]
                    for field in ['name','description','category','sub_category','manufacturer','unit_of_measurement','is_active']:
                        setattr(itm, field, r[field])
                    to_update.append(itm)
                    row_to_item[id(r)] = itm
                else:
                    to_create.append(EquipmentInventoryItem(
                        tenant=tenant,
                        created_by=request.user,
                        name=r['name'],
                        description=r['description'],
                        material_code=r['material_code'],
                        category=r['category'],
                        sub_category=r['sub_category'],
                        manufacturer=r['manufacturer'],
                        unit_of_measurement=r['unit_of_measurement'],
                        specifications={},
                        is_active=r['is_active'],
                    ))

            with transaction.atomic():
                if to_create:
                    EquipmentInventoryItem.objects.bulk_create(to_create, batch_size=1000)
                    created += len(to_create)
                if to_update:
                    EquipmentInventoryItem.objects.bulk_update(to_update, ['name','description','category','sub_category','manufacturer','unit_of_measurement','is_active'], batch_size=1000)
                    updated += len(to_update)

                # Refresh items created to attach technologies
                if to_create:
                    # Fetch back by material codes for new ones
                    new_codes = [i.material_code for i in to_create if i.material_code]
                    if new_codes:
                        for itm in EquipmentInventoryItem.objects.filter(tenant=tenant, material_code__in=new_codes):
                            existing_by_code[itm.material_code] = itm

                # Apply technologies
                for r in rows:
                    tech_list = r['technologies']
                    if not tech_list:
                        continue
                    item = None
                    if r['material_code'] and r['material_code'] in existing_by_code:
                        item = existing_by_code[r['material_code']]
                    else:
                        # If no code, find by unique tuple created earlier
                        item = EquipmentInventoryItem.objects.filter(tenant=tenant, name=r['name'], manufacturer=r['manufacturer'], category=r['category']).first()
                    if item:
                        item.technologies.set([name_to_tag[n] for n in tech_list if n in name_to_tag])

            return Response({'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors, 'sheet_used': (ws.title if 'ws' in locals() and ws is not None else None)})

        return Response({'detail': 'Provide either items[] JSON or file'}, status=400)


